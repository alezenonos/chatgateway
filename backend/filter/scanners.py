import csv
import io
from openpyxl import load_workbook


def extract_text_from_csv(content: bytes) -> str:
    text_parts = []
    decoded = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(decoded))
    for row in reader:
        text_parts.extend(row)
    return " ".join(text_parts)


def extract_text_from_xlsx(content: bytes) -> str:
    text_parts = []
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    text_parts.append(str(cell))
    wb.close()
    return " ".join(text_parts)
